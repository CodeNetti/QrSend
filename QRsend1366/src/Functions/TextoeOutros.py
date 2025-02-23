
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage  # Para manipulação da imagem
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import urllib
from selenium.webdriver.common.by import By
import time
import datetime
import urllib.parse
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QInputDialog, QTextEdit, QApplication, QDialog, QPushButton, QVBoxLayout
from PyQt5.QtGui import QIcon
import pyautogui
from selenium import webdriver
import time
import numpy as np
import os
from Functions.Funcoesdeclique import localizar_imagem_e_clicar, erro_encontrar, aguardar
import pyperclip




usuario = os.getlogin()



def mostrar_opcoes():
    global caminho_envio
 
    global caminho_operacao 
    msg3 = QMessageBox()
    msg3.setIcon(QMessageBox.Question)
    msg3.setWindowTitle("Confirmação")
    msg3.setText("Maravilha!\nVocê deseja encaminhar algum documento após o texto ou acoplar o texto ao documento em uma unica mensagem?")
    btn_acoplado = msg3.addButton("Acoplar Texto", QMessageBox.ActionRole)
    btn_separar = msg3.addButton("Texto Desacoplado", QMessageBox.ActionRole)
    msg3.exec_()
    if msg3.clickedButton() == btn_separar:
        msg4 = QMessageBox()
        msg4.setIcon(QMessageBox.Question)
        msg4.setWindowTitle("Confirmação")
        msg4.setText("Qual o tipo de documento que deseja encaminhar")   
        btn_imagens = msg4.addButton("Imagem", QMessageBox.ActionRole)
        btn_videos = msg4.addButton("Vídeo", QMessageBox.ActionRole)
        btn_documentos = msg4.addButton("Documento", QMessageBox.ActionRole)
        msg4.exec_()
        if msg4.clickedButton() == btn_imagens:
            caminho_operacao = f'../Ver/fotovidio.png'
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha a imagem")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de imagem", "", "Imagens (*.png *.jpg *.jpeg *.bmp)")
            caminho_envio = f'"{caminho_envio}"'
            caminho_envio = caminho_envio.replace("/", "\\")

        elif msg4.clickedButton() == btn_videos:
            caminho_operacao = f'../Ver/fotovidio.png'
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha um vídeo")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de vídeo", "", "Vídeos (*.mp4)")

            if caminho_envio:  # Verifica se um arquivo foi selecionado
                tamanho = False
                while not tamanho:
                    # Remove aspas antes de calcular o tamanho
                    caminho_envio_sem_aspas = caminho_envio.strip('"')

                    # Obtém o tamanho do arquivo
                    tamanho_em_bytes = os.path.getsize(caminho_envio_sem_aspas)
                    tamanho_em_mb = tamanho_em_bytes / (1024 * 1024)  # Converte para MB

                    if tamanho_em_mb >= 63:
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Warning)
                        msg.setText(f"O arquivo selecionado tem {tamanho_em_mb:.2f} MB, que excede o limite de 63 MB.\nDeseja escolher outro arquivo?")
                        msg.setWindowTitle("Confirmação")
                        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                        resposta = msg.exec_()

                        if resposta == QMessageBox.Yes:
                            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de vídeo", "", "Vídeos (*.mp4)")
                        else:
                            mostrar_opcoes()
                            tamanho = True
                    else:
                        # Caminho formatado para o Windows (barras invertidas)
                        caminho_envio = f'"{caminho_envio}"'.replace("/", "\\")
                        tamanho = True

        elif msg4.clickedButton() == btn_documentos:
            caminho_operacao = f'../Ver/documentos.png'
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha o documento")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o documento", "", "Documentos (*.pdf *.xlsx *.csv)")
            caminho_envio = f'"{caminho_envio}"'
            caminho_envio = caminho_envio.replace("/", "\\")
        return "SemAcoplar"

         
    if msg3.clickedButton() == btn_acoplado:
        msg4 = QMessageBox()
        msg4.setIcon(QMessageBox.Question)
        msg4.setWindowTitle("Confirmação")
        msg4.setText("Qual o tipo de documento que deseja encaminhar")   
        btn_imagens = msg4.addButton("Imagem", QMessageBox.ActionRole)
        btn_videos = msg4.addButton("Vídeo", QMessageBox.ActionRole)
        btn_documentos = msg4.addButton("Documento", QMessageBox.ActionRole)
        msg4.exec_()
        if msg4.clickedButton() == btn_imagens:
            caminho_operacao = f'../Ver/fotovidio.png'
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha a imagem")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de imagem", "", "Imagens (*.png *.jpg *.jpeg *.bmp)")
            caminho_envio = f'"{caminho_envio}"'
            caminho_envio = caminho_envio.replace("/", "\\")

        elif msg4.clickedButton() == btn_videos:
            caminho_operacao = f'../Ver/fotovidio.png'
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha um vídeo")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de vídeo", "", "Vídeos (*.mp4)")

            if caminho_envio:  # Verifica se um arquivo foi selecionado
                tamanho = False
                while not tamanho:
                    # Remove aspas antes de calcular o tamanho
                    caminho_envio_sem_aspas = caminho_envio.strip('"')

                    # Obtém o tamanho do arquivo
                    tamanho_em_bytes = os.path.getsize(caminho_envio_sem_aspas)
                    tamanho_em_mb = tamanho_em_bytes / (1024 * 1024)  # Converte para MB

                    if tamanho_em_mb >= 63:
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Warning)
                        msg.setText(f"O arquivo selecionado tem {tamanho_em_mb:.2f} MB, que excede o limite de 63 MB.\nDeseja escolher outro arquivo?")
                        msg.setWindowTitle("Confirmação")
                        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                        resposta = msg.exec_()

                        if resposta == QMessageBox.Yes:
                            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de vídeo", "", "Vídeos (*.mp4)")
                        else:
                            mostrar_opcoes()
                            tamanho = True
                    else:
                        # Caminho formatado para o Windows (barras invertidas)
                        caminho_envio = f'"{caminho_envio}"'.replace("/", "\\")
                        tamanho = True

        elif msg4.clickedButton() == btn_documentos:
            caminho_operacao = f'../Ver/documentos.png'
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha o documento")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_envio, _ = QFileDialog.getOpenFileName(None, "Selecione o documento", "", "Documentos (*.pdf *.xlsx *.csv)")
            caminho_envio = f'"{caminho_envio}"'
            caminho_envio = caminho_envio.replace("/", "\\")       
        return "Acoplar"
         

         
        
# Chama a função



# ENVIO DE MENSAGENS NOMINAIS E  NÃO NOMIANIS APENAS DE TEXTO.





class InputDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Digitação")
        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText("Digite o Texto de Envio")
        self.setWindowIcon(QIcon(f'../Pictures/Logo.png'))
        self.submit_button = QPushButton('Pronto')
        self.submit_button.clicked.connect(self.accept)
        layout = QVBoxLayout()
        layout.addWidget(self.text_edit)
        layout.addWidget(self.submit_button)
        self.setLayout(layout)

    def get_text(self):
        return self.text_edit.toPlainText()
    

class InputDialog2(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Digitação")
        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText("Digite o Texto de Envio | <<nomeconvidado>> = Nome do Convidado | <<nomeacomphantes>> = Nome do acompanhantes")
        self.setWindowIcon(QIcon(f'../Pictures/Logo.png'))
        self.submit_button = QPushButton('Pronto')
        self.submit_button.clicked.connect(self.accept)
        layout = QVBoxLayout()
        layout.addWidget(self.text_edit)
        layout.addWidget(self.submit_button)
        self.setLayout(layout)

    def get_text(self):
        return self.text_edit.toPlainText()

# Função principal
def Envio_Original_Texto2(Dados_Convidados_Envio):
    
    
   
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Question)
    msg.setText("Olá você deseja enviar mensagens nominais?")
    msg.setWindowTitle("Confirmação")
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    resposta = msg.exec_()
    if mostrar_opcoes() == "SemAcoplar":
        if resposta == QMessageBox.No:
            msg2 = QMessageBox()
            msg2.setIcon(QMessageBox.Information)
            msg2.setWindowTitle("Texto de envio")
            msg2.setText("Por Favor digite o texto que deseja enviar")
            msg2.setStandardButtons(QMessageBox.Ok)
            msg2.exec_()
            dialog = InputDialog()
            if dialog.exec_() == QDialog.Accepted:
                    texto_digitado = dialog.get_text()



            if texto_digitado:



                    caminho_fotoinico = f'../Ver/inicio.png'
                    caminho_fotoerro = f'../Ver/erro.png'
                    caminho_fotoseta = f'../Ver/seta.png'
                    caminho_fotoplus = f'../Ver/plus.png'
                    caminho_fotoevidioedoc = caminho_operacao
                    caminho_fotopesquisa = f'../Ver/pesquisa.png'
                    caminho_fotoabrir = f'../Ver/abrir.png'
                    caminho_fotoseta2 = f'../Ver/seta2.png'
                    pyperclip.copy(caminho_envio)

                    msg4 = QMessageBox()
                    msg4.setIcon(QMessageBox.Information)
                    msg4.setWindowTitle("Confirmação")
                    msg4.setText("Clique ok para iniciar o disparo")
                    msg4.setStandardButtons(QMessageBox.Ok)
                    msg4.exec_()



            listatelefonicadeucerto= []
            listatelefonicadeuerrado= []  
    

            nav = webdriver.Chrome()
            nav.maximize_window()
            nav.get("https://web.whatsapp.com/")

            # Aguarda o usuário escanear o QR Code manualmente
            aguardar(caminho_fotoinico, precisao=0.8, intervalo=2)
            agrupado_telefone = Dados_Convidados_Envio.groupby('Telefone')
            for telefone, grupo in agrupado_telefone:
            # Extrair informações principais do grupo
                 nomes = grupo['Nome Convidado'].unique()
                 nomes_convidados = grupo['Nome Acompanhante'].tolist()
                 texto_digitado_refatorado =  urllib.parse.quote(texto_digitado)
                 listatelefonicadeucerto.append([telefone])             
                 link = f"https://web.whatsapp.com/send?phone={telefone}&text={texto_digitado_refatorado}"    
                # Navegar para o WhatsApp Web
                 nav.get(link)
                 time.sleep(10)  
                 if erro_encontrar(caminho_fotoerro, precisao=0.8) == True:
                        print("Imagem encontrada e clicada!")
                        listatelefonicadeuerrado.append([telefone])             
                 else:
                # Usa OpenCV para localizar e clicar na imagem
                   localizar_imagem_e_clicar(caminho_fotoseta, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotoplus, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotoevidioedoc, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotopesquisa, 0.8)
                   time.sleep(1)  
                   pyautogui.hotkey("ctrl", "v")
                   localizar_imagem_e_clicar(caminho_fotoabrir, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotoseta2 , 0.8)
                 time.sleep(5)
            wb = Workbook()
            planilha = wb.active  
            planilha.append(["Telefones","Telefones Erros"])
            # Preenche as colunas com as listas
            for i in range(max(len(listatelefonicadeucerto), len(listatelefonicadeuerrado))):
                linha = [
                listatelefonicadeucerto[i][0] if i < len(listatelefonicadeucerto) else "",  # Evita IndexError
                listatelefonicadeuerrado[i][0] if i < len(listatelefonicadeuerrado) else ""
                ]
                planilha.append(linha)
# Sa    lva          a planilha
            wb.save("dados.xlsx")
            print("Arquivo Excel criado com sucesso!")

        if resposta == QMessageBox.Yes:
            msg2 = QMessageBox()
            msg2.setIcon(QMessageBox.Information)
            msg2.setWindowTitle("Texto de envio")
            msg2.setText("Por Favor digite o texto que deseja enviar")
            msg2.setStandardButtons(QMessageBox.Ok)
            msg2.exec_()
            dialog = InputDialog2()
            if dialog.exec_() == QDialog.Accepted:
                    texto_digitado = dialog.get_text()
            if texto_digitado:
                               # Exibe uma mensagem de confirmação
                
                    caminho_fotoinico = f'../Ver/inicio.png'
                    caminho_fotoerro = f'../Ver/erro.png'
                    caminho_fotoseta = f'../Ver/seta.png'
                    caminho_fotoplus = f'../Ver/plus.png'
                    caminho_fotoevidioedoc = caminho_operacao
                    caminho_fotopesquisa = f'../Ver/pesquisa.png'
                    caminho_fotoabrir = f'../Ver/abrir.png'
                    caminho_fotoseta2 = f'../Ver/seta2.png'
                    pyperclip.copy(caminho_envio)
                    msg4 = QMessageBox()
                    msg4.setIcon(QMessageBox.Information)
                    msg4.setWindowTitle("Confirmação")
                    msg4.setText("Clique ok para iniciar o disparo")
                    msg4.setStandardButtons(QMessageBox.Ok)
                    msg4.exec_()
        # Ap               onto a iamgem, da seta de envio
            listatelefonicadeucerto2= []
            listatelefonicadeuerrado2= [] 
            # Configuração do Selenium
            nav = webdriver.Chrome()
            nav.maximize_window()
            nav.get("https://web.whatsapp.com/")
            # Aguarda o usuário escanear o QR Code manualmente
            aguardar(caminho_fotoinico, precisao=0.8, intervalo=2)
            agrupado_telefone = Dados_Convidados_Envio.groupby('Telefone')
            for telefone, grupo in agrupado_telefone:
        # Ex                          trair informações principais do grupo
                nomes = grupo['Nome Convidado'].dropna().unique()  # Remove valores nulos
                nomes = [str(nome) for nome in nomes]  # Converte todos os valores para string
                nome = ', '.join(nomes)  # Converte o array em string separada por vírgulas
                nomes_convidados = grupo['Nome Acompanhante'].dropna().tolist() 
        # Fo                    rmataif not nomes_convidados:  # Verifica se a lista está vazia
                if not nomes_convidados:
                     nomes_convidados_str =  "" 
                elif len(nomes_convidados) == 1:
                    nomes_convidados_str = nomes_convidados[0]
                elif len(nomes_convidados) == 2:
                    nomes_convidados_str = ' e '.join(nomes_convidados)
                else:
                    nomes_convidados_str = ', '.join(nomes_convidados[:-1]) + ' e ' + nomes_convidados[-1]
                texto_sub = texto_digitado.replace("<<nomeconvidado>>", nome)
                texto_sub = texto_sub.replace("<<nomeacompanhantes>>", nomes_convidados_str)
                textoFormatado = urllib.parse.quote(texto_sub)
                print(telefone)
                listatelefonicadeucerto2.append([telefone])             
                link = f"https://web.whatsapp.com/send?phone={telefone}&text={textoFormatado}"
                nav.get(link)
                time.sleep(10)  
                if erro_encontrar(caminho_fotoerro, precisao=0.8) == True:
                       print("Imagem encontrada e clicada!")
                       listatelefonicadeuerrado2.append([telefone])             
                else:
                
                  localizar_imagem_e_clicar(caminho_fotoseta, 0.8)
                  time.sleep(4)
                  localizar_imagem_e_clicar(caminho_fotoplus, 0.8)
                  time.sleep(4)
                  localizar_imagem_e_clicar(caminho_fotoevidioedoc, 0.8)
                  time.sleep(4)
                  localizar_imagem_e_clicar(caminho_fotopesquisa, 0.8)
                  time.sleep(1)  
                  pyautogui.hotkey("ctrl", "v")
                  localizar_imagem_e_clicar(caminho_fotoabrir, 0.8)
                  time.sleep(4)
                  localizar_imagem_e_clicar(caminho_fotoseta2 , 0.8)
                time.sleep(5)
            wb = Workbook()
            planilha = wb.active  
            planilha.append(["Telefones","Telefones Erros"])
            # Preenche as colunas com as listas
            for i in range(max(len(listatelefonicadeucerto2), len(listatelefonicadeuerrado2))):
                linha = [
                listatelefonicadeucerto2[i][0] if i < len(listatelefonicadeucerto2) else "",  # Evita IndexError
                listatelefonicadeuerrado2[i][0] if i < len(listatelefonicadeuerrado2) else ""
                ]
                planilha.append(linha)

####################### Verific ar o Tipo de Envio True Or False
####################### Verific ar o Tipo de Envio
####################### Verific ar o Tipo de Envio
####################### Verific ar o Tipo de Envio    
# 
#           
#    
    else:
        if resposta == QMessageBox.No:
            msg2 = QMessageBox()
            msg2.setIcon(QMessageBox.Information)
            msg2.setWindowTitle("Texto de envio")
            msg2.setText("Por Favor digite o texto que deseja enviar")
            msg2.setStandardButtons(QMessageBox.Ok)
            msg2.exec_()
            dialog = InputDialog()
            if dialog.exec_() == QDialog.Accepted:
                        texto_digitado = dialog.get_text()



            if texto_digitado:

                        caminho_fotoinico = f'../Ver/inicio.png'
                        caminho_fotoerro = f'../Ver/erro.png'
                        caminho_fotoseta = f'../Ver/seta.png'
                        caminho_fotoplus = f'../Ver/plus.png'
                        caminho_fotoevidioedoc = caminho_operacao
                        caminho_fotopesquisa = f'../Ver/pesquisa.png'
                        caminho_fotoabrir = f'../Ver/abrir.png'
                        caminho_fotoseta2 = f'../Ver/seta2.png'
                        caminho_textoacoplado = f'../Ver/textoacoplado.png'
                        pyperclip.copy(caminho_envio)


                        msg4 = QMessageBox()
                        msg4.setIcon(QMessageBox.Information)
                        msg4.setWindowTitle("Confirmação")
                        msg4.setText("Clique ok para iniciar o disparo")
                        msg4.setStandardButtons(QMessageBox.Ok)
                        msg4.exec_()



            listatelefonicadeucerto= []
            listatelefonicadeuerrado= [] 
            nav = webdriver.Chrome()
            nav.maximize_window()
            nav.get("https://web.whatsapp.com/")
            # Aguarda o usuário escanear o QR Code manualmente
            aguardar(caminho_fotoinico, precisao=0.8, intervalo=2)
            agrupado_telefone = Dados_Convidados_Envio.groupby('Telefone')
            for telefone, grupo in agrupado_telefone:
            # Extrair informações principais do grupo
                 nomes = grupo['Nome Convidado'].unique()
                 nomes_convidados = grupo['Nome Acompanhante'].tolist()
                 texto_digitado_refatorado =  texto_digitado.encode('utf-8').decode('utf-8')
                 listatelefonicadeucerto.append([telefone])             
                 link = f"https://web.whatsapp.com/send?phone={telefone}"    
                # Navegar para o WhatsApp Web
                 nav.get(link)
                 time.sleep(10)  
                 if erro_encontrar(caminho_fotoerro, precisao=0.8) == True:
                        print("Imagem encontrada e clicada!")
                        listatelefonicadeuerrado.append([telefone])             
                 else:
                # Usa OpenCV para localizar e clicar na imagem
                   localizar_imagem_e_clicar(caminho_fotoseta, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotoplus, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotoevidioedoc, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_fotopesquisa, 0.8)
                   time.sleep(1)  
                   pyautogui.hotkey("ctrl", "v")
                   localizar_imagem_e_clicar(caminho_fotoabrir, 0.8)
                   time.sleep(4)
                   localizar_imagem_e_clicar(caminho_textoacoplado, 0.8)
                   time.sleep(1)
                   pyperclip.copy(texto_digitado_refatorado)
                   pyautogui.hotkey("ctrl", "v")
                   pyperclip.copy(caminho_envio)                 
                   localizar_imagem_e_clicar(caminho_fotoseta2 , 0.8)
                 time.sleep(3)
            wb = Workbook()
            planilha = wb.active  
            planilha.append(["Telefones","Telefones Erros"])
            # Preenche as colunas com as listas
            for i in range(max(len(listatelefonicadeucerto), len(listatelefonicadeuerrado))):
                linha = [
                listatelefonicadeucerto[i][0] if i < len(listatelefonicadeucerto) else "",  # Evita IndexError
                listatelefonicadeuerrado[i][0] if i < len(listatelefonicadeuerrado) else ""
                ]
                planilha.append(linha)
# Salva              a planilha
            wb.save("dados.xlsx")
            print("Arquivo Excel criado com sucesso!")
        if resposta == QMessageBox.Yes:
            msg2 = QMessageBox()
            msg2.setIcon(QMessageBox.Information)
            msg2.setWindowTitle("Texto de envio")
            msg2.setText("Por Favor digite o texto que deseja enviar")
            msg2.setStandardButtons(QMessageBox.Ok)
            msg2.exec_()
            dialog = InputDialog2()
            if dialog.exec_() == QDialog.Accepted:
                    texto_digitado = dialog.get_text()
            if texto_digitado:
                           # Exibe uma mensagem de confirmação
                
                caminho_fotoinico = f'../Ver/inicio.png'
                caminho_fotoerro = f'../Ver/erro.png'
                caminho_fotoseta = f'../Ver/seta.png'
                caminho_fotoplus = f'../Ver/plus.png'
                caminho_fotoevidioedoc = caminho_operacao
                caminho_fotopesquisa = f'../Ver/pesquisa.png'
                caminho_fotoabrir = f'../Ver/abrir.png'
                caminho_fotoseta2 = f'../Ver/seta2.png'
                caminho_textoacoplado = f'../Ver/textoacoplado.png'
                pyperclip.copy(caminho_envio)
        
                       
                msg4 = QMessageBox()
                msg4.setIcon(QMessageBox.Information)
                msg4.setWindowTitle("Confirmação")
                msg4.setText("Clique ok para iniciar o disparo")
                msg4.setStandardButtons(QMessageBox.Ok)
                msg4.exec_()

    # Ap               onto a iamgem, da seta de envio
            listatelefonicadeucerto2= []
            listatelefonicadeuerrado2= [] 

                       # Configuração do Selenium
            nav = webdriver.Chrome()
            nav.maximize_window()
            nav.get("https://web.whatsapp.com/")
            # Aguarda o usuário escanear o QR Code manualmente
            aguardar(caminho_fotoinico, precisao=0.8, intervalo=2)
            agrupado_telefone = Dados_Convidados_Envio.groupby('Telefone')

        
            for telefone, grupo in agrupado_telefone:
    # Ex               trair informações principais do grupo
                            nomes = grupo['Nome Convidado'].dropna().unique()  # Remove valores nulos
                            nomes = [str(nome) for nome in nomes]  # Converte todos os valores para string
                            nome = ', '.join(nomes)  # Converte o array em string separada por vírgulas

                            nomes_convidados = grupo['Nome Acompanhante'].dropna().tolist() 

    # Fo               rmataif not nomes_convidados:  # Verifica se a lista está vazia
                            if not nomes_convidados:
                                 nomes_convidados_str =  "" 
                            elif len(nomes_convidados) == 1:
                                nomes_convidados_str = nomes_convidados[0]
                            elif len(nomes_convidados) == 2:
                                nomes_convidados_str = ' e '.join(nomes_convidados)
                            else:
                                nomes_convidados_str = ', '.join(nomes_convidados[:-1]) + ' e ' + nomes_convidados[-1]

                            texto_sub = texto_digitado.replace("<<nomeconvidado>>", nome)
                            texto_sub = texto_sub.replace("<<nomeacompanhantes>>", nomes_convidados_str)
                            #texto para o formato URL caso precise
                            #textoFormatado = urllib.parse.quote(texto_sub)
                            textoFormatado = texto_sub.encode('utf-8').decode('utf-8')
        
                            print(telefone)
        
                            listatelefonicadeucerto2.append([telefone])             
                            link = f"https://web.whatsapp.com/send?phone={telefone}"

                            nav.get(link)
                            time.sleep(10)  
                            if erro_encontrar(caminho_fotoerro, precisao=0.8) == True:
                                   print("Imagem encontrada e clicada!")
                                   listatelefonicadeuerrado2.append([telefone])             
        
                            else:
                           # Usa OpenCV para localizar e clicar na imagem
                              localizar_imagem_e_clicar(caminho_fotoseta, 0.8)
                              time.sleep(3)
                              localizar_imagem_e_clicar(caminho_fotoplus, 0.8)
                              time.sleep(3)
                              localizar_imagem_e_clicar(caminho_fotoevidioedoc, 0.8)
                              time.sleep(3)
                              localizar_imagem_e_clicar(caminho_fotopesquisa, 0.8)
                              time.sleep(1)  
                              pyautogui.hotkey("ctrl", "v")
                              localizar_imagem_e_clicar(caminho_fotoabrir, 0.8)
                              time.sleep(4)
                              localizar_imagem_e_clicar(caminho_textoacoplado, 0.8)
                              pyperclip.copy(textoFormatado)
                              pyautogui.hotkey("ctrl", "v")
                              time.sleep(2)
                              pyperclip.copy(caminho_envio)     
                              localizar_imagem_e_clicar(caminho_fotoseta2 , 0.8)
                            time.sleep(3)
                            
            wb = Workbook()
            planilha = wb.active  
            planilha.append(["Telefones","Telefones Erros"])
            # Preenche as colunas com as listas
            for i in range(max(len(listatelefonicadeucerto2), len(listatelefonicadeuerrado2))):
                linha = [
                listatelefonicadeucerto2[i][0] if i < len(listatelefonicadeucerto2) else "",  # Evita IndexError
                listatelefonicadeuerrado2[i][0] if i < len(listatelefonicadeuerrado2) else ""
                ]
                planilha.append(linha)
                     
        
# Salva a planilha

    data_atual = datetime.datetime.now()

# Formata a data no formato desejado (por exemplo, "AAAA-MM-DD")
    data_formatada = data_atual.strftime("%Y-%m-%d")

# Define o caminho e o nome do arquivo com a data incluída
    caminho_arquivo = f"../Resultados/dadosenvio_{data_formatada}.xlsx"

    print(planilha)
    wb.save(caminho_arquivo)
    print("Arquivo Excel criado com sucesso!")
