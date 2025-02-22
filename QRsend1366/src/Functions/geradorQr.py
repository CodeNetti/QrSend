
import qrcode
import json
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage, ImageDraw, ImageFont
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import urllib.parse
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import pyautogui
import requests
import os
from PyQt5.QtWidgets import QFileDialog, QMessageBox,QApplication, QLabel,QVBoxLayout,QDialog, QFileDialog, QMessageBox, QPushButton,QInputDialog
from PyQt5.QtGui import QPixmap , QDesktopServices
from PyQt5.QtCore import QUrl
from openpyxl.drawing.image import Image as ExcelImage

caminho_foto_fundo_final =''


def ajustar_imagem_para_celula(img, cell_width, cell_height):
    # Calcula as proporções para redimensionar a imagem dentro da célula
    scale_w = cell_width / img.width
    scale_h = cell_height / img.height
    scale = min(scale_w, scale_h)  # Escolhe o menor fator de escala para manter a proporção
    
    # Aplica a escala
    img.width = int(img.width * scale)
    img.height = int(img.height * scale)
    
    return img

def exibir_imagem(caminho_imagem):
    dialog = QDialog()
    dialog.setWindowTitle("Pré-visualização")

    # Cria um layout vertical
    layout = QVBoxLayout()

    # Cria um QLabel e carrega a imagem
    label = QLabel()
    pixmap = QPixmap(caminho_imagem)

    # Obter a largura e altura originais da imagem
    largura_original = pixmap.width()
    altura_original = pixmap.height()

    # Defina o tamanho desejado de exibição (ajuste se necessário)
    largura_desejada = 400
    proporcao = largura_desejada / largura_original
    altura_desejada = int(altura_original * proporcao)

    # Redimensionar o QPixmap proporcionalmente
    pixmap = pixmap.scaled(largura_desejada, altura_desejada)

    # Configura o pixmap no QLabel
    label.setPixmap(pixmap)

    # Cria um botão "OK" para continuar
    btn_ok = QPushButton("OK")
    btn_ok.clicked.connect(dialog.accept)  # Fecha o diálogo quando o botão é clicado

    # Adiciona o QLabel e o botão ao layout
    layout.addWidget(label)
    layout.addWidget(btn_ok)

    # Configura o layout do diálogo
    dialog.setLayout(layout)
    dialog.exec_()

# Exemplo de uso da função
def Iterar_sobre_as_imagens(PastaQrs, ImagemdeFundo, PastaQrsFundo):
    # Definir o caminho da pasta onde os QR codes com fundo serão salvos
    usuario = os.getlogin()
    pasta_qrs_com_fundo = PastaQrsFundo

    # Limpar a pasta QrsComFundo antes de salvar novas imagens
    for arquivo in os.listdir(pasta_qrs_com_fundo):
        caminho_arquivo = os.path.join(pasta_qrs_com_fundo, arquivo)
        try:
            if os.path.isfile(caminho_arquivo):  # Verifica se é um arquivo
                os.remove(caminho_arquivo)  # Remove o arquivo
        except Exception as e:
            print(f"Erro ao tentar excluir o arquivo {arquivo}: {e}")

    # Iterar sobre os arquivos da pasta de QR codes
    for arquivo in os.listdir(PastaQrs):
        if arquivo.endswith(".png"):  # Verificar se o arquivo é uma imagem PNG
            caminho_qrcode = os.path.join(PastaQrs, arquivo)

            # Abrir o QR code e a imagem de fundo
            qrcode_img = PILImage.open(caminho_qrcode).convert("RGBA")
            fundo_img = PILImage.open(ImagemdeFundo).convert("RGB")

            # Verificar se o QR code é maior que a imagem de fundo
            if qrcode_img.width > fundo_img.width or qrcode_img.height > fundo_img.height:
                # Calcular a escala para redimensionar o QR code
                escala_w = fundo_img.width / qrcode_img.width
                escala_h = fundo_img.height / qrcode_img.height
                escala = min(escala_w, escala_h)

                # Redimensionar o QR code
                novo_tamanho = (
                    int(qrcode_img.width * escala),
                    int(qrcode_img.height * escala),
                )
                qrcode_img = qrcode_img.resize(novo_tamanho, PILImage.Resampling.LANCZOS)

            # Calcular a posição para colar o QR code no centro da imagem de fundo
            pos_x = (fundo_img.width - qrcode_img.width) // 2
            pos_y = (fundo_img.height - qrcode_img.height) // 2

            # Colar o QR code na imagem de fundo
            fundo_img.paste(qrcode_img, (pos_x, pos_y), qrcode_img)

            # Construir o caminho completo com o nome do arquivo e extensão
            caminho_qrcode_salvar = os.path.join(pasta_qrs_com_fundo, arquivo)

            # Salvar a imagem resultante no novo caminho
            fundo_img.save(caminho_qrcode_salvar)


def Carregar_Plano_De_Fundo(usuario):
    global caminho_pasta
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Question)
    msg.setText("Você deseja atribuir uma imagem de fundo aos QrCodes gerados?")
    msg.setWindowTitle("Confirmação")
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    resposta = msg.exec_()

    if resposta == QMessageBox.Yes:
            caminhoQrs = f"../Qrs/"
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Procurar arquivo")
            msg.setText("Escolha a imagem de fundo do QRcode")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            caminho_fotofundo, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de imagem", "", "Imagens (*.png *.jpg *.jpeg *.bmp)")
            caminho_pasta = QFileDialog.getExistingDirectory(None, "Selecione a pasta aonde deseja salvar os QR codes")
            Iterar_sobre_as_imagens(caminhoQrs,caminho_fotofundo,caminho_pasta)
            caminho_foto_fundo_final = f"{caminho_pasta}/qrcode_1.png"
            exibir_imagem(caminho_foto_fundo_final)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setText("Você deseja refazer o fundo?")
            msg.setWindowTitle("Confirmação")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            resposta_fundo = msg.exec_()
            while resposta_fundo == QMessageBox.Yes:
                caminhoQrs = f"../Qrs/"
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Procurar arquivo")
                msg.setText("Escolha a nova imagem de fundo do QRcode")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                caminho_fotofundo, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de imagem", "", "Imagens (*.png *.jpg *.jpeg *.bmp)")
                Iterar_sobre_as_imagens(caminhoQrs,caminho_fotofundo,caminho_pasta)
                exibir_imagem(caminho_foto_fundo_final)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Question)
                msg.setText("Você deseja refazer o fundo?")
                msg.setWindowTitle("Confirmação")
                msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                resposta_fundo = msg.exec_()

            


def criar_Planilha_Final():
    # Criar uma nova planilha do Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dados Convidados'

 

    # Definir os nomes das colunas
    ws['A1'] = 'ID'
    ws['B1'] = 'Nome Convidado'
    ws['C1'] = 'Nome Acompanhante'
    ws['D1'] = 'Nome QRcode'
    ws['E1'] = 'Telefone'
    ws['F1'] = 'QrCode'
    ws['G1'] = 'Faixa Etaria'
    ws['H1'] = 'Idade'

 
    max_rows = 2000 
    max_columns = 6 

    # Definir altura das linhas
    for row in range(2, max_rows + 2):  # Começa a partir da linha 2 para deixar a linha 1 para os cabeçalhos
        ws.row_dimensions[row].height = 150  # Altura em pontos

    # Definir largura das colunas
    for col in range(1, max_columns + 1):
        col_letter = openpyxl.utils.get_column_letter(col)  # Converte índice de coluna para letra
        ws.column_dimensions[col_letter].width = 25  # Largura em caracteres
        
    
    return wb , ws 
    


def insere_Planilha_Final(Dados_Convidados , ws , wb):
    usuario = os.getlogin()
    pasta_qrs = f"../Qrs"

    for arquivo in os.listdir(pasta_qrs):
        caminho_arquivo = os.path.join(pasta_qrs, arquivo)
        try:
            if os.path.isfile(caminho_arquivo):
                os.remove(caminho_arquivo)  # Remove o arquivo
        except Exception as e:
            print(f"Erro ao tentar excluir o arquivo {arquivo}: {e}")
    
    id = 1
    i = 0
    linha_excel = 2  # Começa na segunda linha, logo após os cabeçalhos
    
    #Dados_Convidados = pd.read_excel(dados_Convidados)
    
    font_path = "C:/Windows/Fonts/Gadugi.ttf"
    
    while i < len(Dados_Convidados):
        valorNome = Dados_Convidados.at[i, 'Nome Convidado']
        valorNomeConvidado = str(Dados_Convidados.at[i, 'Nome Acompanhante'])
        valorNomeQrCode = str(Dados_Convidados.at[i, 'Nome QrCode'])
        valorTel = str(Dados_Convidados.at[i, 'Telefone'])  
        valorFaixaEtaria = str(Dados_Convidados.at[i, 'Faixa Etaria']) 
        valorIdade = str(Dados_Convidados.at[i, 'Idade']) 
    
    
        data = {
            "ID": id,
            "Nome Convidado": valorNome,
            "Nome Acompanhante": valorNomeConvidado,
            "Nome QrCode": valorNomeQrCode ,
            "Telefone": valorTel,
            "Faixa Etaria": valorFaixaEtaria,
            "Idade": valorIdade
        }
        print(data)
        id += 1  
        json_data = json.dumps(data)   
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(json.dumps(json_data))
        qr.make(fit=True)
    
        img = qr.make_image(fill_color="black", back_color="white")
    
        font = ImageFont.truetype(font_path,50)
    
        draw = ImageDraw.Draw(img)
        text_bbox = draw.textbbox((0, 0), valorNomeQrCode, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[0]
        qr_width, qr_height = img.size
        new_width = max(qr_width, text_width)
        new_height = qr_height + text_height + 20
    
        text_img = PILImage.new('RGB', (new_width, new_height), 'white')
        text_img.paste(img, ((new_width - qr_width) // 2, 0))
    
        draw = ImageDraw.Draw(text_img)
        text_position = ((new_width - text_width) // 2, qr_height + 10)
        draw.text(text_position, valorNomeQrCode, font=font, fill="black")
    
        usuario = os.getlogin()
        qr_image_path = f"../Qrs/qrcode_{data['ID']}.png"
    
        os.makedirs(os.path.dirname(qr_image_path), exist_ok=True)
    
        text_img.save(qr_image_path)
    
        print(f"QR code Gerado: {qr_image_path}")

        i += 1
        file_saved = False
    while not file_saved:
        try:
            Carregar_Plano_De_Fundo(usuario)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Salvar arquivo")
            msg.setText("Aonde gostaria de salvar o arquivo?")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()  # Incrementa o índice do loop principal
            QFileDialog.Options() 
            filePath, _ = QFileDialog.getSaveFileName(None, "Salvar Planilha", "", "Excel Files (*.xlsx)")
            if not filePath:
                return  # Se o usuário cancelar a operação, sair da função

            # Tenta salvar a planilha
            wb.save(filePath)
            file_saved = True  # Se salvar com sucesso, sai do loop
            
            # Mensagem de confirmação de salvamento
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Confirmação")
            msg.setText("Arquivo salvo com sucesso!")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

        except PermissionError:
            # Exibe uma mensagem de erro se a planilha estiver aberta
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Salvar Planilha")
            msg.setText("Não foi possível salvar a planilha. Certifique-se de que o arquivo não está aberto e tente novamente.")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            file_saved = False  # Se salvar com sucesso, sai do loop
    wb = load_workbook(filePath)
    ws = wb.active
    i = 0
    id = 1
    linha_excel = 2 # Incrementa a linha para o próximo QR code
    usuario = os.getlogin()
    while i < len(Dados_Convidados):
            valorNome = Dados_Convidados.at[i, 'Nome Convidado']
            valorNomeConvidado = str(Dados_Convidados.at[i, 'Nome Acompanhante'])
            valorNomeQrCode = str(Dados_Convidados.at[i, 'Nome QrCode'])
            valorTel = str(Dados_Convidados.at[i, 'Telefone'])  
            valorFaixaEtaria = str(Dados_Convidados.at[i, 'Faixa Etaria']) 
            valorIdade = str(Dados_Convidados.at[i, 'Idade']) 
            data = {
                "ID": id,
                "Nome Convidado": valorNome,
                "Nome Acompanhante": valorNomeConvidado,
                "Nome QrCode": valorNomeQrCode,
                "Telefone": valorTel,
                "Faixa Etaria": valorFaixaEtaria,
                "Idade": valorIdade

            }
            
            qr_image_path2 = f"{caminho_pasta}/qrcode_{data['ID']}.png"
            text_img = ExcelImage(qr_image_path2)

# Defina a largura e altura desejadas (em pixels) que correspondem ao tamanho da célula
            largura_celula = 400  # Ajuste conforme necessário
            altura_celula = 200  # Ajuste conforme necessário
            
            # Redimensiona a imagem
            
            # Adiciona o QR code redimensionado à planilha
            text_img = ajustar_imagem_para_celula(text_img, largura_celula, altura_celula)
            ws[f'A{linha_excel}'] = data["ID"]
            ws[f'B{linha_excel}'] = data["Nome Convidado"]
            ws[f'C{linha_excel}'] = data["Nome Acompanhante"]
            ws[f'D{linha_excel}'] = data["Nome QrCode"]
            ws[f'E{linha_excel}'] = data["Telefone"]
            text_img.anchor = f'F{linha_excel}'
            ws.add_image(text_img)
            ws[f'G{linha_excel}'] = data["Faixa Etaria"]
            ws[f'H{linha_excel}'] = data["Idade"]
            linha_excel += 1  # Incrementa a linha para o próximo QR code
            id += 1 
            i += 1
            wb.save(filePath)
            wb.close()
                        
            

def abrir_Arquivo():
    msg2 = QMessageBox()
    msg2.setIcon(QMessageBox.Question)
    msg2.setText("Você deseja abrir o arquivo?")
    msg2.setWindowTitle("Confirmação")
    msg2.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    respostaOpen = msg2.exec_()  # Chama apenas uma vez
    if respostaOpen == QMessageBox.Yes:
        Arquivo, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo Excel", "", "Excel Files (*.xlsx *.xls)")
        QDesktopServices.openUrl(QUrl.fromLocalFile(Arquivo))
        
    

